from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from core.auth import get_current_user
from core.config import supabase
import logging
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/reports", tags=["Reports"])


# ==========================
# EXPORTAR RELATÓRIO DE TURMA (EXCEL)
# ==========================

@router.get("/class/{class_id}/excel")
def export_class_report_excel(class_id: str, user=Depends(get_current_user)):
    """Exporta relatório da turma em Excel"""
    
    try:
        # Buscar dados da turma
        class_data = supabase.table("classes") \
            .select("name") \
            .eq("id", class_id) \
            .single() \
            .execute()
        
        class_name = class_data.data.get("name", "Turma") if class_data.data else "Turma"
        
        # Buscar alunos
        students = supabase.table("students") \
            .select("id, name, registration_number") \
            .eq("class_id", class_id) \
            .execute()
        
        students_data = students.data or []
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório"
        
        # Cabeçalho
        ws["A1"] = f"Relatório da Turma: {class_name}"
        ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells("A1:F1")
        
        ws["A2"] = f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A2"].font = Font(italic=True, size=10)
        
        # Cabeçalhos das colunas
        headers = ["Aluno", "Matrícula", "Última Nota", "Status", "Tendência", "Submissões"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Dados dos alunos
        row = 5
        for student in students_data:
            # Buscar submissões
            submissions = supabase.table("student_submissions") \
                .select("score, status") \
                .eq("student_id", student["id"]) \
                .eq("class_id", class_id) \
                .order("created_at", desc=True) \
                .execute()
            
            subs = submissions.data or []
            
            if subs:
                latest = subs[0]
                score = latest.get("score", 0)
                status = latest.get("status", "pending")
                
                # Determinar status
                if status == "ausente":
                    status_text = "Ausente"
                elif score >= 6:
                    status_text = "Aprovado"
                elif score >= 5:
                    status_text = "Em Risco"
                else:
                    status_text = "Reprovado"
                
                # Determinar tendência
                if len(subs) >= 2:
                    prev_score = subs[1].get("score", 0)
                    if score > prev_score:
                        trend = "↑ Melhorando"
                    elif score < prev_score:
                        trend = "↓ Piorando"
                    else:
                        trend = "→ Estável"
                else:
                    trend = "—"
                
                # Preencher linha
                ws.cell(row=row, column=1).value = student["name"]
                ws.cell(row=row, column=2).value = student["registration_number"]
                ws.cell(row=row, column=3).value = score
                ws.cell(row=row, column=4).value = status_text
                ws.cell(row=row, column=5).value = trend
                ws.cell(row=row, column=6).value = len(subs)
                
                # Colorir status
                status_cell = ws.cell(row=row, column=4)
                if status_text == "Aprovado":
                    status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    status_cell.font = Font(color="006100")
                elif status_text == "Em Risco":
                    status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    status_cell.font = Font(color="9C6500")
                elif status_text == "Reprovado":
                    status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    status_cell.font = Font(color="9C0006")
                
                row += 1
        
        # Ajustar largura das colunas
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 12
        
        # Converter para bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=relatorio_turma_{class_name}.xlsx"}
        )
    
    except Exception as e:
        logger.error(f"Erro ao exportar relatório: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao exportar: {str(e)}")


# ==========================
# EXPORTAR RELATÓRIO DE ALUNO (EXCEL)
# ==========================

@router.get("/student/{student_id}/excel")
def export_student_report_excel(student_id: str, user=Depends(get_current_user)):
    """Exporta relatório individual do aluno em Excel"""
    
    try:
        # Buscar dados do aluno
        student = supabase.table("students") \
            .select("name, registration_number") \
            .eq("id", student_id) \
            .single() \
            .execute()
        
        student_name = student.data.get("name", "Aluno") if student.data else "Aluno"
        student_reg = student.data.get("registration_number", "—") if student.data else "—"
        
        # Buscar submissões
        submissions = supabase.table("student_submissions") \
            .select("score, status, created_at, assessment_id") \
            .eq("student_id", student_id) \
            .order("created_at") \
            .execute()
        
        subs = submissions.data or []
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório"
        
        # Cabeçalho
        ws["A1"] = f"Relatório do Aluno: {student_name}"
        ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells("A1:D1")
        
        ws["A2"] = f"Matrícula: {student_reg}"
        ws["A2"].font = Font(italic=True, size=10)
        
        ws["A3"] = f"Data do Relatório: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A3"].font = Font(italic=True, size=10)
        
        # Cabeçalhos das colunas
        headers = ["Data", "Avaliação", "Nota", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Dados das submissões
        row = 6
        scores = []
        for sub in subs:
            if sub.get("status") == "corrected":
                score = sub.get("score", 0)
                scores.append(score)
                
                ws.cell(row=row, column=1).value = datetime.fromisoformat(sub["created_at"]).strftime("%d/%m/%Y")
                ws.cell(row=row, column=2).value = sub.get("assessment_id", "—")
                ws.cell(row=row, column=3).value = score
                
                if score >= 6:
                    status_text = "Aprovado"
                    color = "C6EFCE"
                    font_color = "006100"
                elif score >= 5:
                    status_text = "Em Risco"
                    color = "FFEB9C"
                    font_color = "9C6500"
                else:
                    status_text = "Reprovado"
                    color = "FFC7CE"
                    font_color = "9C0006"
                
                status_cell = ws.cell(row=row, column=4)
                status_cell.value = status_text
                status_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                status_cell.font = Font(color=font_color)
                
                row += 1
        
        # Resumo
        if scores:
            avg_score = sum(scores) / len(scores)
            ws[f"A{row+1}"] = "Média:"
            ws[f"A{row+1}"].font = Font(bold=True)
            ws[f"B{row+1}"] = avg_score
            ws[f"B{row+1}"].font = Font(bold=True)
        
        # Ajustar largura das colunas
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 12
        
        # Converter para bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=relatorio_aluno_{student_name}.xlsx"}
        )
    
    except Exception as e:
        logger.error(f"Erro ao exportar relatório: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao exportar: {str(e)}")


# ==========================
# EXPORTAR RESUMO DA ESCOLA (EXCEL)
# ==========================

@router.get("/school/summary/excel")
def export_school_summary_excel(user=Depends(get_current_user)):
    """Exporta resumo geral da escola em Excel"""
    
    if user["role"] not in ("admin", "super_admin"):
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    try:
        school_id = user["school_id"]
        
        # Buscar turmas
        classes = supabase.table("classes") \
            .select("id, name") \
            .eq("school_id", school_id) \
            .execute()
        
        classes_data = classes.data or []
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumo"
        
        # Cabeçalho
        ws["A1"] = "Resumo Geral da Escola"
        ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells("A1:F1")
        
        ws["A2"] = f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A2"].font = Font(italic=True, size=10)
        
        # Cabeçalhos
        headers = ["Turma", "Média", "Taxa Aprovação", "Aprovados", "Reprovados", "Total"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Dados das turmas
        row = 5
        for cls in classes_data:
            submissions = supabase.table("student_submissions") \
                .select("score, status") \
                .eq("class_id", cls["id"]) \
                .execute()
            
            subs = submissions.data or []
            scores = [s.get("score", 0) for s in subs if s.get("status") == "corrected" and s.get("score") is not None]
            
            average = sum(scores) / len(scores) if scores else 0
            approved = sum(1 for s in subs if s.get("status") == "corrected" and s.get("score", 0) >= 6)
            failed = sum(1 for s in subs if s.get("status") == "corrected" and s.get("score", 0) < 6)
            approval_rate = (approved / (approved + failed) * 100) if (approved + failed) > 0 else 0
            
            ws.cell(row=row, column=1).value = cls["name"]
            ws.cell(row=row, column=2).value = round(average, 2)
            ws.cell(row=row, column=3).value = f"{round(approval_rate, 1)}%"
            ws.cell(row=row, column=4).value = approved
            ws.cell(row=row, column=5).value = failed
            ws.cell(row=row, column=6).value = len(subs)
            
            row += 1
        
        # Ajustar largura
        for col in range(1, 7):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        # Converter para bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=resumo_escola.xlsx"}
        )
    
    except Exception as e:
        logger.error(f"Erro ao exportar resumo: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao exportar: {str(e)}")
