from fastapi import FastAPI, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, Response
from pydantic import BaseModel
import requests
import os
import io
import pandas as pd

from core.config import supabase
from core.auth import get_current_user

from routers import schools
from routers import classes
from routers import students
from routers import assessments
from routers import ocr
from routers import manual
from routers import teachers
from routers import dashboard
from routers import users
from routers import admin_dashboard
from routers import teacher_dashboard
from routers import reports
from routers import alerts
from routers.subjects import router as subjects_router

app = FastAPI(title="Calia Digital API", version="2.0.0")

# ==========================
# CORS — permitir frontend antigo, novo e localhost
# ==========================

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://localhost:5173",
        "https://calia-frontend.vercel.app",
        "https://calia-frontend-v2.vercel.app",
        "https://caliadigital.com.br",
        "https://www.caliadigital.com.br",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==========================
# ROUTERS
# ==========================

app.include_router(schools.router)  # já tem prefix="/schools" no router
app.include_router(classes.router, prefix="/classes")
app.include_router(students.router, prefix="/students")
app.include_router(assessments.router, prefix="/assessments")
app.include_router(ocr.router, prefix="/ocr")
app.include_router(manual.router)
app.include_router(teachers.router, prefix="/teachers")
app.include_router(users.router, prefix="/users")
app.include_router(subjects_router, prefix="/subjects")
app.include_router(dashboard.router)
app.include_router(admin_dashboard.router)
app.include_router(teacher_dashboard.router)
app.include_router(reports.router)
app.include_router(alerts.router)


# ==========================
# MODELS
# ==========================

class PasswordChangeRequest(BaseModel):
    current_password: str
    new_password: str


class ProfileUpdateRequest(BaseModel):
    name: str = None
    full_name: str = None
    avatar_url: str = None
    phone: str = None


# ==========================
# ENDPOINT /me — Retorna perfil completo do usuário logado
# ==========================

@app.get("/me")
def get_me(user=Depends(get_current_user)):
    """
    Retorna o perfil do usuário logado.
    O frontend usa isso para determinar o role e renderizar o dashboard correto.
    """
    return {
        "id": user.get("id"),
        "email": user.get("email", ""),
        "role": user.get("role", ""),
        "school_id": user.get("school_id"),
        "name": user.get("full_name") or user.get("name") or user.get("email", ""),
        "full_name": user.get("full_name", ""),
    }


# ==========================
# ENDPOINT /me/change-password
# ==========================

@app.put("/me/change-password")
def change_password(data: PasswordChangeRequest, user=Depends(get_current_user)):
    """
    Permite que o usuário logado altere sua senha.
    Requer a senha atual para validação.
    """
    
    if not data.current_password or not data.new_password:
        raise HTTPException(status_code=400, detail="Senha atual e nova senha são obrigatórias")
    
    if len(data.new_password) < 6:
        raise HTTPException(status_code=400, detail="Nova senha deve ter pelo menos 6 caracteres")
    
    try:
        # Configurações do Supabase
        SUPABASE_URL = os.getenv("SUPABASE_URL", "https://lhydfllckxuzotondmla.supabase.co")
        SUPABASE_ANON_KEY = os.getenv("SUPABASE_ANON_KEY", "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImxoeWRmbGxja3h1em90b25kbWxhIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzIzOTYxNTQsImV4cCI6MjA4Nzk3MjE1NH0.PvwKkuSX8tJXHmoztSodHqMoFCsbJyslhDHnxeAGHjs")
        
        # Validar senha atual
        login_response = requests.post(
            f"{SUPABASE_URL}/auth/v1/token?grant_type=password",
            json={"email": user["email"], "password": data.current_password},
            headers={
                "Content-Type": "application/json",
                "apikey": SUPABASE_ANON_KEY,
            }
        )
        
        if not login_response.ok:
            raise HTTPException(status_code=401, detail="Senha atual incorreta")
        
        # Alterar senha usando o endpoint do Supabase
        access_token = login_response.json().get("access_token")
        update_response = requests.put(
            f"{SUPABASE_URL}/auth/v1/user",
            json={"password": data.new_password},
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {access_token}",
                "apikey": SUPABASE_ANON_KEY,
            }
        )
        
        if not update_response.ok:
            raise HTTPException(status_code=500, detail="Erro ao alterar senha")
        
        return {"message": "Senha alterada com sucesso"}
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao alterar senha: {str(e)}")


# ==========================
# ENDPOINT /me/update
# ==========================

@app.put("/me/update")
def update_profile(data: ProfileUpdateRequest, user=Depends(get_current_user)):
    """
    Permite que o usuário logado atualize seu perfil.
    """
    
    try:
        # Preparar dados para atualizar
        update_data = {}
        if data.name:
            update_data["full_name"] = data.name
        if data.full_name:
            update_data["full_name"] = data.full_name
        if data.avatar_url:
            update_data["avatar_url"] = data.avatar_url
        if data.phone:
            update_data["phone"] = data.phone
        
        if not update_data:
            raise HTTPException(status_code=400, detail="Nenhum campo para atualizar")
        
        # Atualizar na tabela profiles
        result = supabase.table("profiles") \
            .update(update_data) \
            .eq("id", user["id"]) \
            .execute()
        
        if not result.data:
            raise HTTPException(status_code=500, detail="Erro ao atualizar perfil")
        
        return {"message": "Perfil atualizado com sucesso", "data": result.data[0]}
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao atualizar perfil: {str(e)}")


# ==========================
# HEALTH CHECK
# ==========================

# ==========================
# ENDPOINTS PARA DOWNLOAD DE MODELOS
# ==========================

@app.get("/templates/teachers")
def download_teachers_template(user=Depends(get_current_user)):
    """
    Retorna um arquivo Excel com o modelo para importar professores em lote.
    Colunas: Nome Completo, Email
    """
    if user["role"] not in ("admin", "super_admin"):
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Professores"
    
    # Cabeçalho
    headers = ["Nome Completo", "Email"]
    ws.append(headers)
    
    # Formatação do cabeçalho
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # Adicionar exemplos
    examples = [
        ["João Silva", "joao.silva@email.com"],
        ["Maria Santos", "maria.santos@email.com"]
    ]
    
    for row in examples:
        ws.append(row)
    
    # Formatação das linhas de exemplo
    for row in ws.iter_rows(min_row=2, max_row=len(examples)+1, min_col=1, max_col=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Ajustar largura das colunas
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    
    # Converter para bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=modelo_professores.xlsx"}
    )


@app.get("/templates/students")
def download_students_template(user=Depends(get_current_user)):
    """
    Retorna um arquivo Excel com o modelo para importar alunos em lote.
    Colunas: Nome, Matricula
    """
    if user["role"] not in ("admin", "super_admin"):
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Alunos"
    
    # Cabeçalho
    headers = ["Nome", "Matricula"]
    ws.append(headers)
    
    # Formatação do cabeçalho
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # Adicionar exemplos
    examples = [
        ["Ana Silva", "2024001"],
        ["Bruno Costa", "2024002"]
    ]
    
    for row in examples:
        ws.append(row)
    
    # Formatação das linhas de exemplo
    for row in ws.iter_rows(min_row=2, max_row=len(examples)+1, min_col=1, max_col=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Ajustar largura das colunas
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20
    
    # Converter para bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=modelo_alunos.xlsx"}
    )


@app.get("/")
def root():
    return {"status": "CALIA backend online", "version": "2.0.0"}
